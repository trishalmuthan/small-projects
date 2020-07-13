import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Function process_workbook takes filename and for every row in the 3nd column (starting from the 2nd row because of the label),
# will reduce price by 10% and writes the new price in the 4th column. Then, a bar graph is created storing all the values in the 4th column

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        og_cell = sheet.cell(row, 3)
        corrected_price = og_cell.value * 0.9
        new_corrected_cell = sheet.cell(row, 4)
        new_corrected_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
